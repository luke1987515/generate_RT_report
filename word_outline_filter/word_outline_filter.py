import os
import sys
import win32com.client
from openpyxl import Workbook, load_workbook
import argparse

# Word Constants
wdOutlineLevelBodyText = 10

def get_word_app():
    try:
        word = win32com.client.GetActiveObject("Word.Application")
    except:
        word = win32com.client.Dispatch("Word.Application")
    word.Visible = False
    return word

def export_to_excel(word_file_path, excel_file_path):
    print(f"[*] Opening Word document: {word_file_path}")
    word = get_word_app()
    abs_word_path = os.path.abspath(word_file_path)
    doc = word.Documents.Open(abs_word_path)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Outline"
    ws.append(["Index", "Level", "Title", "Status (Keep/Delete)"])
    
    print("[*] Scanning paragraphs for outline levels...")
    count = 0
    total_paras = doc.Paragraphs.Count
    for i in range(1, total_paras + 1):
        try:
            p = doc.Paragraphs(i)
            level = p.OutlineLevel
            if level < wdOutlineLevelBodyText:
                # Skip TOC entries by checking Style name
                style_name = ""
                try:
                    style_name = p.Style.NameLocal
                except:
                    pass
                
                if "TOC" in style_name.upper() or "目錄" in style_name:
                    continue

                text = p.Range.Text.strip()
                # Only include non-empty titles or titles with some text
                if text:
                    ws.append([i, level, text, "Keep"])
                    count += 1
        except Exception as e:
            # Skip paragraphs that might cause issues (e.g. deleted/special objects)
            continue
    
    wb.save(excel_file_path)
    doc.Close()
    print(f"[+] Exported {count} headings to {excel_file_path}")

def apply_filter(word_file_path, excel_file_path, output_file_path):
    print(f"[*] Reading filter rules from: {excel_file_path}")
    wb = load_workbook(excel_file_path)
    ws = wb.active
    
    print(f"[*] Analyzing deletion ranges and propagating up...")
    
    # Map for quick level and status lookup
    idx_to_level = {row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    all_headings = sorted(idx_to_level.keys())
    
    # 1. Initial deleted set from Excel
    deleted_indices = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        idx, level, title, status = row
        if idx and status and status.strip().lower() == "delete":
            deleted_indices.add(idx)

    # 2. Downward propagation: If a parent is deleted, all its descendants are also effectively deleted
    explicit_deletes = list(deleted_indices)
    for idx in explicit_deletes:
        level = idx_to_level[idx]
        for h_idx in all_headings:
            if h_idx > idx:
                if idx_to_level[h_idx] <= level:
                    break
                deleted_indices.add(h_idx)

    # 3. Upward propagation: If all children are deleted, mark parent for deletion
    # We run this in multiple passes to handle nesting (L4->L3 then L3->L2)
    for pass_num in range(2):
        changed = False
        for idx in all_headings:
            if idx in deleted_indices:
                continue
            
            level = idx_to_level[idx]
            if level not in [2, 3]: 
                continue
            
            # Find all headings that belong to this parent
            children = []
            has_sub_headings = False
            for h_idx in all_headings:
                if h_idx > idx:
                    h_level = idx_to_level[h_idx]
                    if h_level <= level:
                        break
                    has_sub_headings = True
                    children.append(h_idx)
            
            # If it has sub-headings and ALL of them are now in deleted_indices
            if has_sub_headings and all(c in deleted_indices for c in children):
                deleted_indices.add(idx)
                title = [row[2] for row in ws.iter_rows(min_row=2, values_only=True) if row[0] == idx][0]
                print(f"[i] Auto-marking empty parent for deletion: Level {level} '{title}'")
                changed = True
        
        if not changed:
            break

    # 4. Identify final intended deletion ranges [start_idx, end_idx)
    raw_ranges = []
    for idx in deleted_indices:
        level = idx_to_level[idx]
        # Find boundary: next heading with level <= current level
        end_idx = None
        for h_idx in all_headings:
            if h_idx > idx and idx_to_level[h_idx] <= level:
                end_idx = h_idx
                break
        raw_ranges.append([idx, end_idx])

    if not raw_ranges:
        print("[!] No sections marked for deletion. Exiting.")
        return

    # 2. Merge overlapping ranges (e.g. if both parent and child are marked for delete)
    raw_ranges.sort()
    merged_ranges = []
    if raw_ranges:
        curr_start, curr_end = raw_ranges[0]
        for next_start, next_end in raw_ranges[1:]:
            if curr_end is None or next_start < curr_end:
                if next_end is None:
                    curr_end = None
                elif curr_end is not None and next_end > curr_end:
                    curr_end = next_end
            else:
                merged_ranges.append((curr_start, curr_end))
                curr_start, curr_end = next_start, next_end
        merged_ranges.append((curr_start, curr_end))

    print(f"[*] Merged into {len(merged_ranges)} deletion blocks.")
    
    print(f"[*] Opening Word document: {word_file_path}")
    word = get_word_app()
    abs_word_path = os.path.abspath(word_file_path)
    doc = word.Documents.Open(abs_word_path)
    
    # 3. Delete from bottom to top to maintain index validity for the start of each block
    merged_ranges.sort(key=lambda x: x[0], reverse=True)
    for start_idx, end_idx in merged_ranges:
        try:
            # Note: Word Paragraphs are 1-indexed.
            target_para = doc.Paragraphs(start_idx)
            start_pos = target_para.Range.Start
            
            if end_idx is None:
                end_pos = doc.Range().End
            else:
                # Find the current index of the boundary paragraph
                # Since we are going backwards, the start_idx is stable.
                # However, the end_idx might have changed if we deleted something AFTER it?
                # Wait, if we are going backwards, we delete things at the END first.
                # The indices of paragraphs BEFORE the deletion point are UNCHANGED.
                # So doc.Paragraphs(start_idx) is correct.
                # And since we merged ranges, there are no deletions happening BETWEEN start_idx and end_idx
                # in a way that would shift end_idx within the current step.
                end_para = doc.Paragraphs(end_idx)
                end_pos = end_para.Range.Start
            
            title_text = target_para.Range.Text.strip()[:30]
            doc.Range(start_pos, end_pos).Delete()
            print(f"[-] Deleted block starting at index {start_idx} ('{title_text}')")
        except Exception as e:
            print(f"[!] Error deleting block at index {start_idx}: {e}")

    print(f"[*] Updating Table of Contents...")
    try:
        # Disable alerts to prevent "Update Table of Contents" dialog
        word.DisplayAlerts = 0 # wdAlertsNone
        for toc in doc.TablesOfContents:
            toc.Update()
        print("[+] Table of Contents updated.")
    except Exception as e:
        print(f"[!] Could not update TOC: {e}")
    finally:
        word.DisplayAlerts = -1 # wdAlertsAll

    abs_output_path = os.path.abspath(output_file_path)
    wdFormatXMLDocument = 12
    try:
        doc.SaveAs(abs_output_path, FileFormat=wdFormatXMLDocument)
    except:
        doc.SaveAs(abs_output_path)
        
    doc.Close()
    print(f"[+] Filtered document saved to: {output_file_path}")

def main():
    parser = argparse.ArgumentParser(description="Word Outline Filter Tool")
    parser.add_argument("--export", help="Export Word outline to Excel", action="store_true")
    parser.add_argument("--apply", help="Apply Excel filter to Word document", action="store_true")
    parser.add_argument("--input", help="Input Word file", default="Template.doc")
    parser.add_argument("--excel", help="Excel file path", default="outline.xlsx")
    parser.add_argument("--output", help="Output Word file", default="Filtered_Template.docx")
    
    args = parser.parse_args()

    if args.export:
        export_to_excel(args.input, args.excel)
    elif args.apply:
        apply_filter(args.input, args.excel, args.output)
    else:
        # Interactive mode if no flags
        print("=== Word Outline Filter Tool ===")
        print("1. Export Outline to Excel")
        print("2. Apply Filter from Excel to Word")
        choice = input("Select an option (1/2): ")
        
        if choice == "1":
            word_in = input(f"Input Word path [{args.input}]: ") or args.input
            excel_out = input(f"Output Excel path [{args.excel}]: ") or args.excel
            export_to_excel(word_in, excel_out)
        elif choice == "2":
            word_in = input(f"Input Word path [{args.input}]: ") or args.input
            excel_in = input(f"Input Excel path [{args.excel}]: ") or args.excel
            word_out = input(f"Output Word path [{args.output}]: ") or args.output
            apply_filter(word_in, excel_in, word_out)
        else:
            print("Invalid choice.")

if __name__ == "__main__":
    main()
